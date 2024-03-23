import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

class DataEntryApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Data Entry App")
        self.create_widgets()

    def create_widgets(self):
        # Labels
        tk.Label(self.master, text="Name:").grid(row=0)
        tk.Label(self.master, text="Age:").grid(row=1)

        # Entry widgets
        self.name_entry = tk.Entry(self.master)
        self.age_entry = tk.Entry(self.master)

        self.name_entry.grid(row=0, column=1, padx=10, pady=10)
        self.age_entry.grid(row=1, column=1, padx=10, pady=10)

        # Submit button
        tk.Button(self.master, text="Submit", command=self.submit_data).grid(row=2, column=0, columnspan=2, pady=10)

    def submit_data(self):
        name = self.name_entry.get()
        age = self.age_entry.get()

        if name and age:
            # Append data to Excel
            self.append_to_excel(name, age)
            # Clear entry fields
            self.name_entry.delete(0, tk.END)
            self.age_entry.delete(0, tk.END)
        else:
            tk.messagebox.showwarning("Error", "Please fill in all fields.")

    def append_to_excel(self, name, age):
        workbook_path = "data_entry.xlsx"
        try:
            workbook = load_workbook(workbook_path)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.remove(workbook.active)
            workbook.create_sheet("Data", 0)

        sheet = workbook["Data"]
        sheet.append([name, age])

        workbook.save(workbook_path)
        tk.messagebox.showinfo("Success", "Data added to Excel.")

if __name__ == "__main__":
    root = tk.Tk()
    app = DataEntryApp(root)
    root.mainloop()